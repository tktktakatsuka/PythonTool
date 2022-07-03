
import os
import re
# resuests モジュールをインポート
from ast import IsNot, Str
from contextlib import nullcontext
from operator import truediv
from bs4.builder import HTML
import openpyxl
from openpyxl.xml.constants import ACTIVEX, XLSX
from openpyxl.styles import PatternFill
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.utils import keys_to_typing
from selenium.webdriver.chrome.options import Options
import time
import datetime
from selenium.webdriver.common.action_chains import ActionChains
import chromedriver_binary
import lxml.html
import logging
import random
import configparser
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains


def click_by_position(driver, x, y) -> None:
    from selenium.webdriver.common.action_chains import ActionChains
    actions = ActionChains(driver)

    # MOVE TO TOP_LEFT (`move_to_element` will guide you to the CENTER of the element)
    whole_page = driver.find_element_by_xpath('//*[@id="myAccountPopupOrders"]')
    actions.move_to_element_with_offset(whole_page, 0, 0)

    # MOVE TO DESIRED POSITION THEN CLICK
    actions.move_by_offset(x, y)
    actions.click()

    actions.perform()


click_by_position()


"""#フィールドを定義する

"""

config_ini = configparser.ConfigParser()
base = os.path.dirname(os.path.abspath(__file__))


config_ini.read('config.ini', encoding='utf-8')
var2 = config_ini.get('DEFAULT', 'Driverpath')
var3 = config_ini.get('DEFAULT', 'InputWb')


inputWb = openpyxl.load_workbook(var3)
inputWs = inputWb.worksheets[0]
reserchkaisu = inputWs.cell(row = 3, column = 11).value
max = inputWs.max_row
options = Options()

user_agent = ['Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36',
                  'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36',
                  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.157 Safari/537.36',
                  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36'
                  ]
options = webdriver.ChromeOptions()
options.add_argument('--user-agent=' + user_agent[random.randrange(0, len(user_agent), 1)])
#options.add_argument('--headless')
driver = webdriver.Chromedriver = webdriver.Chrome(executable_path= var2 , chrome_options=options)

URL ="https://www.costco.co.jp/login"
time.sleep(2)
#URLを開く
driver.get(URL)
time.sleep(3)
#カレントページを取得する

#メールアドレスを記入
xpath = '//*[@id="j_username"]'
elem_login_btn = driver.find_element_by_xpath(xpath)
elem_login_btn.send_keys('tomoko.tomoko.0909@gmail.com')

#パスワードを記入
xpath = '//*[@id="j_password"]'
elem_login_btn = driver.find_element_by_xpath(xpath)
elem_login_btn.send_keys('costco1qaz')

#loginする
xpath = '//*[@id="loginSubmit"]'
elem_login_btn = driver.find_element_by_xpath(xpath)
elem_login_btn.click()



#loginの座f票を取得する。
time.sleep(5)
xpath = '//*[@id="account-menu-trigger"]'
elem_login_btn = driver.find_element_by_xpath(xpath)
print(elem_login_btn.location)
ActionChains.move_by_offset(elem_login_btn.location)
